# debt_manager_excel_sync.py
# Purpose: Synchronizes data between the SQLite database and the Excel dashboard.
#          Uses pandas for efficient data transfer and openpyxl for Excel interaction.
# Deploy in: C:\DebtTracker
# Version: 1.1 (2025-07-18) - Initial version for Python-based synchronization.

import os
import sqlite3
import pandas as pd
import logging
import xlwings as xw # Using xlwings for checking if Excel is open
from openpyxl import load_workbook, Workbook # Using openpyxl for data read/write to sheets

# Import configuration from config.py
from config import DB_PATH, EXCEL_PATH, LOG_FILE, LOG_DIR, TABLE_SCHEMAS

# Ensure log directory exists
os.makedirs(LOG_DIR, exist_ok=True)

# Configure logging
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s: %(message)s',
                    handlers=[
                        logging.FileHandler(LOG_FILE, mode='a'),
                        logging.StreamHandler()
                    ])

def get_db_connection():
    """Establishes and returns a SQLite database connection."""
    conn = None
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row # Allows accessing columns by name
        return conn
    except sqlite3.Error as e:
        logging.error(f"Database connection error: {e}")
        # Do not show messagebox here, let the caller handle it
        return None

def sync_data(direction):
    """
    Synchronizes data between SQLite database and Excel dashboard.
    Args:
        direction (str): 'sqlite_to_excel' or 'excel_to_sqlite'.
    """
    logging.info(f"Starting {direction} sync...")

    # Check if Excel is open using xlwings (more reliable than process check)
    try:
        app = xw.App(visible=False) # Try to get an app instance
        if app.books: # If there are any open workbooks
            # Check if our specific Excel file is open
            is_our_excel_open = False
            for book in app.books:
                if os.path.normpath(book.fullname) == os.path.normpath(EXCEL_PATH):
                    is_our_excel_open = True
                    break

            if is_our_excel_open:
                logging.warning('Excel file is open, skipping sync to prevent data corruption.')
                # Messagebox should be handled by the GUI caller, not here.
                # If run standalone, it will just log.
                app.quit() # Quit the app instance we created
                return False
        app.quit() # Quit the app instance if no workbooks were open or ours wasn't
    except Exception as e:
        logging.warning(f"Could not check if Excel is open using xlwings: {e}. Proceeding with caution.")
        # If xlwings fails, we still try to proceed with openpyxl, which might fail if file is locked.

    if not os.path.exists(EXCEL_PATH):
        logging.error(f"Excel file not found: {EXCEL_PATH}. Cannot perform sync.")
        return False

    conn = None
    try:
        conn = get_db_connection()
        if conn is None:
            return False

        # Load all data from SQLite into DataFrames
        sqlite_dfs = {}
        for table_name in TABLE_SCHEMAS.keys():
            try:
                # Use a direct SQL query to load, ensuring all columns are fetched
                # This handles cases where schema might have evolved in DB but not yet in code
                df = pd.read_sql_query(f"SELECT * FROM {table_name}", conn)
                # Ensure numeric columns are properly typed. This is crucial as SQLite is typeless.
                numeric_cols = [
                    'Amount', 'OriginalAmount', 'AmountPaid', 'MinimumPayment', 'SnowballPayment',
                    'InterestRate', 'Balance', 'StartingBalance', 'PreviousBalance', 'Value', 'TargetAmount',
                    'CurrentAmount', 'AllocationPercentage', 'NextProjectedIncome', 'AccountLimit'
                ]
                for col in numeric_cols:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)

                # Explicitly convert ID columns to integer, filling NaNs with 0
                id_cols = ['DebtID', 'AccountID', 'PaymentID', 'GoalID', 'AssetID', 'RevenueID', 'CategoryID', 'AllocatedTo']
                for col in id_cols:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

                sqlite_dfs[table_name] = df
            except pd.io.sql.DatabaseError as e:
                logging.warning(f"Table '{table_name}' not found in SQLite or other DB error: {e}. Skipping load from SQLite.")
                sqlite_dfs[table_name] = pd.DataFrame(columns=TABLE_SCHEMAS[table_name]['db_columns']) # Provide empty DF

        # Load all data from Excel into DataFrames
        excel_dfs = {}
        try:
            excel_workbook = load_workbook(EXCEL_PATH)
            for table_name in TABLE_SCHEMAS.keys():
                if table_name not in excel_workbook.sheetnames:
                    logging.warning(f"Sheet '{table_name}' not found in Excel. Skipping load from Excel.")
                    excel_dfs[table_name] = pd.DataFrame(columns=TABLE_SCHEMAS[table_name]['excel_columns'])
                    continue

                sheet = excel_workbook[table_name]
                # Read data starting from the second row (skipping headers)
                data = sheet.values
                # Get headers from the first row
                excel_headers = [cell.value for cell in sheet[1]]

                # Create DataFrame, skipping the header row from data
                df = pd.DataFrame(data=list(data)[1:], columns=excel_headers)
                excel_dfs[table_name] = df
        except Exception as e:
            logging.error(f"Error loading data from Excel: {e}")
            return False

        if direction == 'sqlite_to_excel':
            for table_name, df_sqlite in sqlite_dfs.items():
                if table_name not in excel_workbook.sheetnames:
                    logging.warning(f"Sheet '{table_name}' not found in Excel for write. Skipping {table_name} sync.")
                    continue

                sheet = excel_workbook[table_name]
                # Clear existing data, but keep headers (first row)
                if sheet.max_row > 1:
                    sheet.delete_rows(2, sheet.max_row)

                # Prepare data for writing to Excel, ensuring column order matches Excel headers
                excel_columns_for_table = TABLE_SCHEMAS[table_name]['excel_columns']
                # Special handling for Debts to include 'Projected Payment' in Excel
                if table_name == 'Debts' and 'Projected Payment' not in excel_columns_for_table:
                    excel_columns_for_table = excel_columns_for_table + ['Projected Payment']
                    # Calculate Projected Payment if not already in df_sqlite (for display only)
                    if 'ProjectedPayment' not in df_sqlite.columns:
                        df_sqlite['ProjectedPayment'] = df_sqlite['MinimumPayment'] + df_sqlite['SnowballPayment']

                # Map DB column names to Excel column names for writing
                db_to_excel_map = {db_col: excel_col for db_col, excel_col in zip(TABLE_SCHEMAS[table_name]['db_columns'], TABLE_SCHEMAS[table_name]['excel_columns'])}

                # Create a DataFrame with Excel column names and order
                df_to_write = pd.DataFrame()
                for db_col, excel_col in db_to_excel_map.items():
                    if db_col in df_sqlite.columns:
                        df_to_write[excel_col] = df_sqlite[db_col]

                # Add 'Projected Payment' if it's a Debts table and was calculated
                if table_name == 'Debts' and 'Projected Payment' in excel_columns_for_table and 'ProjectedPayment' in df_sqlite.columns:
                    df_to_write['Projected Payment'] = df_sqlite['ProjectedPayment'].apply(lambda x: f"${x:,.2f}") # Format for Excel display

                # Ensure all expected Excel columns are present, even if empty in DB
                for col_name in excel_columns_for_table:
                    if col_name not in df_to_write.columns:
                        df_to_write[col_name] = None # Add missing columns

                # Reorder columns to match Excel headers
                df_to_write = df_to_write[excel_columns_for_table]

                # Append data to sheet (starting from row 2)
                for r_idx, row in df_to_write.iterrows():
                    sheet.append(list(row.values))
                logging.info(f"Synced data from SQLite table '{table_name}' to Excel sheet '{table_name}'.")

            excel_workbook.save(EXCEL_PATH)
            logging.info(f"sqlite_to_excel sync completed successfully.")

        elif direction == 'excel_to_sqlite':
            for table_name, df_excel in excel_dfs.items():
                if table_name not in TABLE_SCHEMAS:
                    logging.warning(f"Table '{table_name}' from Excel not defined in schema. Skipping sync to SQLite.")
                    continue

                # Prepare DataFrame for writing to SQLite, ensuring column order matches DB schema
                db_columns_for_table = TABLE_SCHEMAS[table_name]['db_columns']
                excel_columns_for_table = TABLE_SCHEMAS[table_name]['excel_columns']

                # Create a mapping from Excel column names to DB column names
                excel_to_db_map = {excel_col: db_col for db_col, excel_col in zip(db_columns_for_table, excel_columns_for_table)}

                df_to_write = pd.DataFrame()
                for excel_col, db_col in excel_to_db_map.items():
                    if excel_col in df_excel.columns:
                        # Clean up Excel-specific formatting (e.g., '$', '%') before saving to DB
                        if df_excel[excel_col].dtype == 'object': # If it's a string
                            if 'Amount' in db_col or 'Payment' in db_col or 'Balance' in db_col or 'Value' in db_col or 'Income' in db_col or 'Limit' in db_col:
                                df_to_write[db_col] = df_excel[excel_col].astype(str).str.replace(r'[$,]', '', regex=True).astype(float)
                            elif 'InterestRate' in db_col:
                                df_to_write[db_col] = df_excel[excel_col].astype(str).str.replace(r'%', '', regex=True).astype(float)
                            elif 'Date' in db_col or 'Received' in db_col:
                                # Attempt to parse date strings into datetime objects, then format for SQLite
                                df_to_write[db_col] = pd.to_datetime(df_excel[excel_col], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S')
                            else:
                                df_to_write[db_col] = df_excel[excel_col]
                        else:
                            df_to_write[db_col] = df_excel[excel_col]
                    else:
                        df_to_write[db_col] = None # Add missing columns as None

                # Ensure all expected DB columns are present and in the correct order
                df_to_write = df_to_write[db_columns_for_table]

                # Write to SQLite
                conn.execute(f"DELETE FROM {table_name}") # Clear existing data
                df_to_write.to_sql(table_name, conn, if_exists='append', index=False)
                logging.info(f"Synced data from Excel sheet '{table_name}' to SQLite table '{table_name}'.")

            conn.commit()
            logging.info(f"excel_to_sqlite sync completed successfully.")

        else:
            logging.error(f"Invalid sync direction: {direction}")
            return False

    except Exception as e:
        logging.error(f"Error during {direction} sync: {e}", exc_info=True)
        return False
    finally:
        if conn:
            conn.close()
            logging.info("SQLite connection closed after sync.")

if __name__ == "__main__":
    # This block is for testing the script independently.
    # When run via the GUI or orchestrator, sync_data will be called directly.
    print("This script is primarily for synchronization logic.")
    print("To test, ensure DebtManager.db and DebtDashboard.xlsx exist.")
    print("Example: sync_data('sqlite_to_excel') or sync_data('excel_to_sqlite')")
    # Example usage (uncomment to test):
    # try:
    #     sync_data('sqlite_to_excel')
    #     print("SQLite to Excel sync attempted.")
    # except Exception as e:
    #     print(f"Sync failed: {e}")
