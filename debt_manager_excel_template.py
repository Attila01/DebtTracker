# debt_manager_excel_template.py
# Purpose: Creates or updates the Excel dashboard template (DebtDashboard.xlsx)
#          NOTE: This script is now considered legacy as the main app uses CSV sync.
# Deploy in: C:\DebtTracker
# Version: 2.5 (2025-07-21) - Corrected table re-creation logic to prevent errors.
#                            - Removed all dependencies on xlwings, now uses openpyxl exclusively.

import os
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from config import EXCEL_PATH, TABLE_SCHEMAS, LOG_FILE, LOG_DIR

os.makedirs(LOG_DIR, exist_ok=True)

if not logging.getLogger().handlers:
    logging.basicConfig(level=logging.INFO,
                        format='%(asctime)s: %(message)s',
                        handlers=[
                            logging.FileHandler(LOG_FILE, mode='a'),
                            logging.StreamHandler()
                        ])

def create_excel_template():
    logging.info("Starting Excel template creation/update process.")
    try:
        wb = load_workbook(EXCEL_PATH) if os.path.exists(EXCEL_PATH) else Workbook()

        if 'Dashboard' not in wb.sheetnames:
            dashboard_sheet = wb.create_sheet("Dashboard", 0)
        else:
            dashboard_sheet = wb['Dashboard']
            wb.move_sheet(dashboard_sheet, offset=-wb.sheetnames.index('Dashboard'))

        # Clear dashboard content
        for merged_range in list(dashboard_sheet.merged_cells):
            dashboard_sheet.unmerge_cells(str(merged_range))
        for row in dashboard_sheet.iter_rows():
            for cell in row:
                cell.value = None
                cell.style = 'Normal'

        logging.info("Prepared workbook: ensured 'Dashboard' sheet is first and cleared its content.")

        for table_name, schema in TABLE_SCHEMAS.items():
            if table_name not in wb.sheetnames:
                ws = wb.create_sheet(table_name)
            else:
                ws = wb[table_name]

            # Set headers
            headers = schema['csv_columns']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # --- FIX: Properly clear and recreate tables ---
            # Remove any existing table with the same name from the worksheet
            tables_to_remove = [table.name for table in ws._tables]
            for table_name_to_remove in tables_to_remove:
                try:
                    del ws._tables[table_name_to_remove]
                except KeyError:
                    logging.warning(f"Could not remove table object '{table_name_to_remove}'.")

            # Add new table definition
            table_display_name = f"{table_name}Table"
            tab_ref = f"A1:{get_column_letter(len(headers))}{ws.max_row if ws.max_row > 1 else 1}"
            table = Table(displayName=table_display_name, ref=tab_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            table.tableStyleInfo = style
            ws.add_table(table)
            logging.info(f"Added/updated table '{table_display_name}' on sheet '{table_name}'.")

        # Visual button creation on Dashboard... (logic as provided previously)

        wb.save(EXCEL_PATH)
        logging.info(f"Excel template saved successfully to: {EXCEL_PATH}")

    except Exception as e:
        logging.error(f"Error during Excel template creation: {e}", exc_info=True)
        raise

if __name__ == "__main__":
    try:
        create_excel_template()
        print(f"Excel template created/updated at: {EXCEL_PATH}.")
    except Exception as e:
        print(f"Failed to create/update Excel template: {e}.")